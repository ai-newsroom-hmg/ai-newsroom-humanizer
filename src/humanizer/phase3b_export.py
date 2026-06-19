"""Phase 3b Excel + Demo-Export.

Inputs:
  data/phase3b/nightly_results.jsonl   — alle Varianten (variant-level)
  data/phase3b/nightly_summary.json    — per-Doc-Summary
  data/phase3b/nightly_target_final.txt — Best-Variant des Demo-Texts
  data/phase3b/pangram_pre_n13.jsonl   — Pre-Scores (Wackel-Test)

Outputs:
  ~/Downloads/humanize-phase3b/results.xlsx
    Sheet "Übersicht"     — Casdorff- + OOD-Aggregate, Demo-Target-Status
    Sheet "Pre-Post-Docs" — pro Doc Pre/Post-Pangram, Best-Variante, Faithful-Stats
    Sheet "Variants"      — alle Varianten mit Temp/BGE/Pangram (für Deep-Dive)
    Sheet "Demo-Target"   — original vs final side-by-side
  ~/Downloads/humanize-phase3b/best-variants/<doc_id>.txt
    Pro Doc die humanized Best-Variante als TXT (Demo-fähig)
"""
from __future__ import annotations

import json
import re
from collections import defaultdict
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

ROOT = Path(__file__).resolve().parents[2]
NIGHTLY_VARIANTS = ROOT / "data" / "phase3b" / "nightly_results.jsonl"
NIGHTLY_SUMMARY = ROOT / "data" / "phase3b" / "nightly_summary.json"
TARGET_FINAL = ROOT / "data" / "phase3b" / "nightly_target_final.txt"
TARGET_TRACE = ROOT / "data" / "phase3b" / "nightly_target_trace.json"
TARGET_ORIG = ROOT / "data" / "test-corpora" / "target-text-2026-06-18-journalismus-tot.txt"
PANGRAM_PRE = ROOT / "data" / "phase3b" / "pangram_pre_n13.jsonl"
EVAL = ROOT / "data" / "phase2-training-pool" / "eval.jsonl"

OUT_DIR = Path.home() / "Downloads" / "humanize-phase3b"
OUT_XLSX = OUT_DIR / "results.xlsx"
OUT_TXT_DIR = OUT_DIR / "best-variants"


def load_jsonl(fp: Path) -> list:
    if not fp.exists():
        return []
    # NICHT splitlines() — splittet auch auf U+2028/U+2029, die in Mistral-Outputs
    # vorkommen können (Token-Artefakte). Explizit auf \n.
    return [json.loads(l) for l in fp.read_text(encoding="utf-8").split("\n") if l.strip()]


def safe_fname(s: str, maxlen: int = 60) -> str:
    s = re.sub(r"[^\w\-_.]", "_", s, flags=re.UNICODE)
    return s.strip("_")[:maxlen] or "untitled"


def pick_best_variant(variants: list[dict]) -> dict | None:
    """Best = lowest pangram_fraction_ai under faithful; fallback highest bge."""
    bypass = [v for v in variants if v.get("faithful")
              and v.get("pangram_fraction_ai") is not None
              and v["pangram_fraction_ai"] < 0.2]
    if bypass:
        return min(bypass, key=lambda v: v["pangram_fraction_ai"])
    faith = [v for v in variants if v.get("faithful")
             and v.get("pangram_fraction_ai") is not None]
    if faith:
        return min(faith, key=lambda v: v["pangram_fraction_ai"])
    ok = [v for v in variants if not v.get("error") and v.get("text")]
    if ok:
        return max(ok, key=lambda v: v.get("bge_sim", 0.0))
    return None


def main():
    OUT_DIR.mkdir(parents=True, exist_ok=True)
    OUT_TXT_DIR.mkdir(parents=True, exist_ok=True)

    variants = load_jsonl(NIGHTLY_VARIANTS)
    summary = json.loads(NIGHTLY_SUMMARY.read_text()) if NIGHTLY_SUMMARY.exists() else {}
    pre_rows = {r["doc_id"]: r for r in load_jsonl(PANGRAM_PRE)}
    eval_docs = {d["doc_id"]: d for d in load_jsonl(EVAL)}

    # Variants per doc
    by_doc: dict[str, list[dict]] = defaultdict(list)
    for v in variants:
        by_doc[v["doc_id"]].append(v)

    wb = Workbook()
    HEAD = Font(bold=True, size=11, color="FFFFFF")
    FILL = PatternFill("solid", fgColor="2C3E50")

    # ============ Sheet 1: Übersicht ============
    ws = wb.active
    ws.title = "Übersicht"
    ws["A1"] = "humanize v0.2 — Phase 3b Production-Run (2026-06-19)"
    ws["A1"].font = Font(bold=True, size=14)
    ws["A2"] = "Mistral-3.2-24b-instruct + Best-of-N + BGE-Filter ≥0.85 + Pangram-Rank"
    ws["A2"].font = Font(italic=True, color="555555")

    rows_by_kind: dict[str, list] = defaultdict(list)
    for did, s in summary.items():
        if did == "__TARGET__":
            continue
        rows_by_kind[s["kind"]].append(s)

    row = 5
    for kind in ("casdorff", "ood"):
        rows_k = rows_by_kind.get(kind, [])
        if not rows_k:
            continue
        n_total = len(rows_k)
        n_bypass = sum(1 for s in rows_k if s.get("n_bypass", 0) >= 1)
        n_lt05 = sum(1 for s in rows_k
                     if isinstance(s.get("best_pangram_fraction_ai"), (int, float))
                     and s["best_pangram_fraction_ai"] < 0.5)
        n_faith_avg = (sum(s.get("n_faithful", 0) for s in rows_k) / max(n_total, 1))
        cost = sum(s.get("gen_cost_usd", 0) for s in rows_k)

        ws.cell(row, 1, f"{kind.upper()} Best-of-50/30").font = Font(bold=True, size=12)
        row += 1
        for label, val in (
            ("Docs gesamt", n_total),
            ("Bypass (P<0.2)", f"{n_bypass}/{n_total} ({n_bypass / n_total * 100:.0f}%)"),
            ("Pangram-Post unter 0.5", f"{n_lt05}/{n_total}"),
            ("Ø Faithful-Variants pro Doc", f"{n_faith_avg:.1f}"),
            ("Gen-Kosten", f"${cost:.4f}"),
        ):
            ws.cell(row, 1, label).font = Font(bold=True)
            ws.cell(row, 2, val)
            row += 1
        row += 1

    # Target
    t = summary.get("__TARGET__")
    if t:
        ws.cell(row, 1, "DEMO-TARGET (8675 chars, chunked Best-of-24)").font = Font(bold=True, size=12)
        row += 1
        for label, val in (
            ("Original chars", t.get("orig_chars")),
            ("Humanized chars", t.get("final_chars")),
            ("BGE-Sim global (Multi-Chunk-Min)", round(t.get("bge_sim_global", 0), 4)),
            ("Chunks gesamt", len(t.get("chunk_results", []))),
            ("Gen-Variants total", t.get("n_generated")),
            ("Faithful Variants total", t.get("n_faithful")),
            ("Cost", f"${t.get('cost_usd', 0):.3f}"),
        ):
            ws.cell(row, 1, label).font = Font(bold=True)
            ws.cell(row, 2, val)
            row += 1

    ws.column_dimensions["A"].width = 38
    ws.column_dimensions["B"].width = 26

    # ============ Sheet 2: Pre-Post-Docs ============
    ws2 = wb.create_sheet("Pre-Post-Docs")
    cols = ["#", "kind", "doc_id", "autor", "chars", "n_faithful", "n_bypass",
            "pangram_PRE", "pangram_POST_BEST", "Drop", "best_bge_sim",
            "stopped_reason", "gen_cost"]
    for i, c in enumerate(cols, 1):
        cell = ws2.cell(1, i, c)
        cell.font = HEAD
        cell.fill = FILL
        cell.alignment = Alignment(horizontal="center")

    rowi = 2
    all_rows = [(did, s) for did, s in summary.items() if did != "__TARGET__"]
    all_rows.sort(key=lambda kv: (kv[1].get("kind"),
                                  kv[1].get("best_pangram_fraction_ai") or 1.0))
    for i, (did, s) in enumerate(all_rows, 1):
        d = eval_docs.get(did, {})
        pre_row = pre_rows.get(did, {})
        pre = pre_row.get("pangram_pre_fraction_ai")
        post = s.get("best_pangram_fraction_ai")
        drop = (pre - post) if (isinstance(pre, (int, float))
                                and isinstance(post, (int, float))) else None
        ws2.cell(rowi, 1, i)
        ws2.cell(rowi, 2, s.get("kind"))
        ws2.cell(rowi, 3, did)
        ws2.cell(rowi, 4, s.get("autor") or d.get("autor"))
        ws2.cell(rowi, 5, s.get("chars"))
        ws2.cell(rowi, 6, s.get("n_faithful"))
        ws2.cell(rowi, 7, s.get("n_bypass"))
        ws2.cell(rowi, 8, round(pre, 3) if isinstance(pre, (int, float)) else "n/a")
        ws2.cell(rowi, 9, round(post, 3) if isinstance(post, (int, float)) else "n/a")
        ws2.cell(rowi, 10, round(drop, 3) if isinstance(drop, (int, float)) else "")
        ws2.cell(rowi, 11, round(s.get("best_bge_sim") or 0, 3))
        ws2.cell(rowi, 12, s.get("stopped_reason"))
        ws2.cell(rowi, 13, round(s.get("gen_cost_usd", 0), 4))

        # Color coding
        if s.get("n_bypass", 0) >= 1:
            color = "D4EFDF"  # grün
        elif isinstance(post, (int, float)) and post < 0.5:
            color = "FCEABF"  # gelb
        else:
            color = "FCD7D2"  # rot
        for ci in range(1, len(cols) + 1):
            ws2.cell(rowi, ci).fill = PatternFill("solid", fgColor=color)
        rowi += 1

    for i, w in enumerate([4, 10, 36, 22, 7, 10, 8, 12, 14, 8, 10, 32, 10], 1):
        ws2.column_dimensions[get_column_letter(i)].width = w
    ws2.freeze_panes = "A2"
    ws2.auto_filter.ref = f"A1:{get_column_letter(len(cols))}{rowi - 1}"

    # ============ Sheet 3: Variants ============
    ws3 = wb.create_sheet("Variants")
    vcols = ["doc_id", "kind", "temp", "idx", "bge_sim", "faithful",
             "pangram_fraction_ai", "pangram_prediction", "chars_out", "cost"]
    for i, c in enumerate(vcols, 1):
        cell = ws3.cell(1, i, c)
        cell.font = HEAD
        cell.fill = FILL
    rowi = 2
    for v in variants:
        if v.get("error"):
            continue
        ws3.cell(rowi, 1, v.get("doc_id"))
        ws3.cell(rowi, 2, v.get("kind"))
        ws3.cell(rowi, 3, v.get("temp"))
        ws3.cell(rowi, 4, v.get("idx"))
        ws3.cell(rowi, 5, round(v.get("bge_sim", 0), 4))
        ws3.cell(rowi, 6, "JA" if v.get("faithful") else "nein")
        ws3.cell(rowi, 7, v.get("pangram_fraction_ai"))
        ws3.cell(rowi, 8, v.get("pangram_prediction"))
        ws3.cell(rowi, 9, len(v.get("text", "")))
        ws3.cell(rowi, 10, round(v.get("cost", 0), 5))
        rowi += 1
    for i, w in enumerate([36, 10, 7, 6, 10, 9, 16, 14, 10, 10], 1):
        ws3.column_dimensions[get_column_letter(i)].width = w
    ws3.freeze_panes = "A2"
    ws3.auto_filter.ref = f"A1:{get_column_letter(len(vcols))}{rowi - 1}"

    # ============ Sheet 4: Demo-Target ============
    if TARGET_FINAL.exists() and TARGET_ORIG.exists():
        ws4 = wb.create_sheet("Demo-Target")
        ws4["A1"] = "Demo-Text — Pre/Post — „Von heute an ist der Journalismus tot"
        ws4["A1"].font = Font(bold=True, size=12)
        ws4["A3"] = "PRE (Original)"
        ws4["B3"] = "POST (humanisiert)"
        for c in (ws4["A3"], ws4["B3"]):
            c.font = HEAD
            c.fill = FILL
            c.alignment = Alignment(horizontal="center")
        ws4["A4"] = TARGET_ORIG.read_text()
        ws4["B4"] = TARGET_FINAL.read_text()
        ws4["A4"].alignment = Alignment(wrap_text=True, vertical="top")
        ws4["B4"].alignment = Alignment(wrap_text=True, vertical="top")
        ws4.column_dimensions["A"].width = 70
        ws4.column_dimensions["B"].width = 70
        ws4.row_dimensions[4].height = 600
        if TARGET_TRACE.exists():
            tt = json.loads(TARGET_TRACE.read_text())
            ws4["A6"] = "BGE-Sim global (Multi-Chunk-Min): " + str(round(tt.get("bge_sim_global", 0), 4))
            chunk_lines = ["Chunks:"]
            for c in tt.get("chunk_results", []):
                chunk_lines.append(
                    f"  {c.get('idx')}: {c.get('chars')}->{c.get('out_chars')} chars · "
                    f"bge={c.get('bge_sim')} · pangram={c.get('pangram_post')}"
                )
            ws4["A7"] = "\n".join(chunk_lines)
            ws4["A7"].alignment = Alignment(wrap_text=True, vertical="top")

    # ============ TXT-Export der Best-Variants ============
    for did, s in summary.items():
        if did == "__TARGET__":
            continue
        vs = by_doc.get(did, [])
        best = pick_best_variant(vs)
        if best and best.get("text"):
            d = eval_docs.get(did, {})
            autor = (s.get("autor") or d.get("autor") or "x")
            fname = f"{safe_fname(autor, 30)}__{did[-12:]}.txt"
            header = (
                f"# {autor} | {did}\n"
                f"# kind={s.get('kind')} chars={s.get('chars')} "
                f"n_faithful={s.get('n_faithful')} n_bypass={s.get('n_bypass')}\n"
                f"# pangram_post={best.get('pangram_fraction_ai')} "
                f"bge_sim={best.get('bge_sim')} temp={best.get('temp')}\n\n"
            )
            (OUT_TXT_DIR / fname).write_text(header + best["text"], encoding="utf-8")

    # Demo-Target
    if TARGET_FINAL.exists():
        t = summary.get("__TARGET__") or {}
        header = (
            f"# DEMO-TARGET „Von heute an ist der Journalismus tot\"\n"
            f"# orig_chars={t.get('orig_chars')} final_chars={t.get('final_chars')}\n"
            f"# bge_sim_global={t.get('bge_sim_global')} n_generated={t.get('n_generated')} "
            f"n_faithful={t.get('n_faithful')}\n\n"
        )
        (OUT_TXT_DIR / "DEMO_TARGET_journalismus_tot.txt").write_text(
            header + TARGET_FINAL.read_text(), encoding="utf-8")

    wb.save(OUT_XLSX)
    print(f"--- Excel: {OUT_XLSX}")
    print(f"--- TXT best-variants: {OUT_TXT_DIR}/ ({sum(1 for _ in OUT_TXT_DIR.iterdir())} files)")

    # Final stats
    print("\n=== Phase 3b Aggregat ===")
    for kind, rows_k in rows_by_kind.items():
        n_bp = sum(1 for s in rows_k if s.get("n_bypass", 0) >= 1)
        print(f"  {kind}: Doc-Bypass {n_bp}/{len(rows_k)} = {n_bp / max(len(rows_k), 1) * 100:.0f}%")
    if t := summary.get("__TARGET__"):
        bge = t.get("bge_sim_global", 0)
        print(f"  TARGET: bge_global={bge:.3f} chunks={len(t.get('chunk_results', []))}")


if __name__ == "__main__":
    main()
