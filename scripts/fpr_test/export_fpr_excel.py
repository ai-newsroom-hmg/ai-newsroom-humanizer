"""Excel-Export der FPR-Studie auf pre-2022-Korpus.

Inputs:
  data/fpr-test/pre2022_corpus.jsonl    — Volltexte
  data/fpr-test/pangram_results.jsonl   — Pangram-Werte

Output:
  ~/Downloads/palimpsest-fpr-test/pre2022_pangram_fpr.xlsx
    Sheet "Übersicht"  — FPR-Aggregate (Hard/Strict/Loose), per Jahr/Quelle/Länge
    Sheet "Pre-Post"   — pro Doc: fraction_ai, prediction, source, date, headline
    Sheet "Volltexte"  — pro Doc Volltext (für manuelle Inspektion)
    Sheet "Worst-FP"   — Top-Verdächtige (höchste fraction_ai trotz pre-2022 human)
"""
from __future__ import annotations

import json
from collections import Counter, defaultdict
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

ROOT = Path(__file__).resolve().parents[2]
CORPUS = ROOT / "data" / "fpr-test" / "pre2022_corpus.jsonl"
RESULTS = ROOT / "data" / "fpr-test" / "pangram_results.jsonl"

OUT_DIR = Path.home() / "Downloads" / "palimpsest-fpr-test"
OUT_XLSX = OUT_DIR / "pre2022_pangram_fpr.xlsx"

HEAD = Font(bold=True, size=11, color="FFFFFF")
FILL_HEAD = PatternFill("solid", fgColor="2C3E50")
FILL_HUMAN = PatternFill("solid", fgColor="D4EFDF")
FILL_MIXED_HUMAN = PatternFill("solid", fgColor="FCEABF")
FILL_MIXED_AI = PatternFill("solid", fgColor="F5C16F")
FILL_AI = PatternFill("solid", fgColor="FCD7D2")


def cat_of(fa: float | None) -> str:
    if fa is None:
        return "n/a"
    if fa >= 0.8:
        return "AI"
    if fa >= 0.5:
        return "Mixed-AI"
    if fa >= 0.2:
        return "Mixed-Human"
    return "Human"


def fill_for(fa: float | None):
    c = cat_of(fa)
    return {
        "Human": FILL_HUMAN, "Mixed-Human": FILL_MIXED_HUMAN,
        "Mixed-AI": FILL_MIXED_AI, "AI": FILL_AI,
    }.get(c)


def load_jsonl(fp: Path) -> list[dict]:
    if not fp.exists():
        return []
    return [json.loads(l) for l in fp.read_text(encoding="utf-8").split("\n") if l.strip()]


def main():
    OUT_DIR.mkdir(parents=True, exist_ok=True)
    corpus = {d["doc_id"]: d for d in load_jsonl(CORPUS)}
    results = {r["doc_id"]: r for r in load_jsonl(RESULTS)}
    if not results:
        print(f"ERROR: {RESULTS} fehlt — eval_pangram_fpr.py zuerst laufen lassen.")
        return
    # Merged rows
    rows = []
    for did, d in corpus.items():
        r = results.get(did, {})
        rows.append({
            **d,
            "fraction_ai": r.get("fraction_ai"),
            "fraction_ai_assisted": r.get("fraction_ai_assisted"),
            "fraction_human": r.get("fraction_human"),
            "prediction": r.get("prediction", ""),
            "error": r.get("error", ""),
        })

    ok = [r for r in rows if isinstance(r.get("fraction_ai"), (int, float))]

    wb = Workbook()

    # ============ Sheet 1: Übersicht ============
    ws = wb.active
    ws.title = "Übersicht"
    ws["A1"] = "Palimpsest FPR-Studie: Pangram auf pre-ChatGPT deutschen Leitmedien"
    ws["A1"].font = Font(bold=True, size=14)
    ws["A2"] = f"n={len(ok)} Texte aus 2019-01-01 bis 2021-12-31, alle Tier-1 Print + Online"
    ws["A2"].font = Font(italic=True, color="555555")
    ws["A3"] = "Pangram claimed (Tech Report 2024): ≤ 0.5 %; UChicago Booth 2026 (n=1992): ≤ 1 %"
    ws["A3"].font = Font(italic=True, color="555555")

    row = 5
    ws.cell(row, 1, "False-Positive-Raten").font = Font(bold=True, size=12)
    row += 1
    fp_hard = sum(1 for r in ok if r["fraction_ai"] >= 0.8)
    fp_strict = sum(1 for r in ok if r["fraction_ai"] >= 0.5)
    fp_loose = sum(1 for r in ok if r["fraction_ai"] >= 0.2)
    for label, n in (
        ("Hard FP (fraction_ai >= 0.8) - eindeutig 'AI'", fp_hard),
        ("Strict FP (>= 0.5) - 'Mixed-AI' oder härter", fp_strict),
        ("Loose FP (>= 0.2) - verdächtig ab 'Mixed-Human'", fp_loose),
    ):
        ws.cell(row, 1, label).font = Font(bold=True)
        ws.cell(row, 2, f"{n}/{len(ok)} = {n / max(len(ok), 1) * 100:.1f} %")
        row += 1
    row += 1

    # Per Jahr
    ws.cell(row, 1, "Per Jahr").font = Font(bold=True, size=12)
    row += 1
    by_year = defaultdict(list)
    for r in ok:
        by_year[r["pub_date"][:4]].append(r)
    for y in sorted(by_year):
        rs = by_year[y]
        fp = sum(1 for r in rs if r["fraction_ai"] >= 0.5)
        ws.cell(row, 1, y).font = Font(bold=True)
        ws.cell(row, 2, f"Strict FP {fp}/{len(rs)} = {fp / len(rs) * 100:.1f} %")
        row += 1
    row += 1

    # Per Source
    ws.cell(row, 1, "Per Quelle").font = Font(bold=True, size=12)
    row += 1
    by_src = defaultdict(list)
    for r in ok:
        by_src[r["source"]].append(r)
    for src, rs in sorted(by_src.items(), key=lambda kv: -len(kv[1])):
        fp = sum(1 for r in rs if r["fraction_ai"] >= 0.5)
        ws.cell(row, 1, src).font = Font(bold=True)
        ws.cell(row, 2, f"n={len(rs):>3} | Strict FP {fp}/{len(rs)} = {fp / len(rs) * 100:.1f} %")
        row += 1
    row += 1

    # Per Länge
    ws.cell(row, 1, "Per Länge").font = Font(bold=True, size=12)
    row += 1
    buckets = {"short (<1500)": [], "mid (1500-3000)": [], "long (>=3000)": []}
    for r in ok:
        if r["chars"] < 1500: buckets["short (<1500)"].append(r)
        elif r["chars"] < 3000: buckets["mid (1500-3000)"].append(r)
        else: buckets["long (>=3000)"].append(r)
    for k, rs in buckets.items():
        if not rs:
            continue
        fp = sum(1 for r in rs if r["fraction_ai"] >= 0.5)
        ws.cell(row, 1, k).font = Font(bold=True)
        ws.cell(row, 2, f"n={len(rs):>3} | Strict FP {fp}/{len(rs)} = {fp / len(rs) * 100:.1f} %")
        row += 1
    row += 1

    # Verteilung der Kategorien
    cats = Counter(cat_of(r["fraction_ai"]) for r in ok)
    ws.cell(row, 1, "Pangram-Kategorien").font = Font(bold=True, size=12)
    row += 1
    for cat in ("Human", "Mixed-Human", "Mixed-AI", "AI"):
        n = cats.get(cat, 0)
        ws.cell(row, 1, cat).font = Font(bold=True)
        ws.cell(row, 2, f"{n}/{len(ok)} = {n / max(len(ok), 1) * 100:.1f} %")
        ws.cell(row, 2).fill = fill_for({"Human": 0.0, "Mixed-Human": 0.3, "Mixed-AI": 0.6, "AI": 0.9}[cat])
        row += 1

    ws.column_dimensions["A"].width = 50
    ws.column_dimensions["B"].width = 30

    # ============ Sheet 2: Pre-Post (alle Docs) ============
    ws2 = wb.create_sheet("Pre-Post")
    cols = ["#", "doc_id", "source", "pub_date", "chars", "fraction_ai",
            "fraction_human", "prediction", "Kategorie", "headline", "ressort", "autor"]
    for i, c in enumerate(cols, 1):
        cell = ws2.cell(1, i, c)
        cell.font = HEAD
        cell.fill = FILL_HEAD
        cell.alignment = Alignment(horizontal="center")

    rowi = 2
    rows_sorted = sorted(ok, key=lambda r: -(r.get("fraction_ai") or 0.0))
    for i, r in enumerate(rows_sorted, 1):
        fa = r["fraction_ai"]
        cat = cat_of(fa)
        ws2.cell(rowi, 1, i)
        ws2.cell(rowi, 2, r["doc_id"])
        ws2.cell(rowi, 3, r.get("source", ""))
        ws2.cell(rowi, 4, r.get("pub_date", ""))
        ws2.cell(rowi, 5, r.get("chars", 0))
        ws2.cell(rowi, 6, round(fa, 4) if fa is not None else "n/a")
        ws2.cell(rowi, 7, round(r.get("fraction_human") or 0, 4))
        ws2.cell(rowi, 8, r.get("prediction", ""))
        ws2.cell(rowi, 9, cat)
        ws2.cell(rowi, 10, r.get("headline", ""))
        ws2.cell(rowi, 11, r.get("ressort", ""))
        ws2.cell(rowi, 12, r.get("autor", ""))
        # Color row
        f = fill_for(fa)
        if f:
            for ci in range(1, len(cols) + 1):
                ws2.cell(rowi, ci).fill = f
        rowi += 1

    for i, w in enumerate([4, 36, 22, 12, 8, 10, 10, 12, 14, 50, 22, 22], 1):
        ws2.column_dimensions[get_column_letter(i)].width = w
    ws2.freeze_panes = "A2"
    ws2.auto_filter.ref = f"A1:{get_column_letter(len(cols))}{rowi - 1}"

    # ============ Sheet 3: Volltexte ============
    ws3 = wb.create_sheet("Volltexte")
    vcols = ["doc_id", "source", "pub_date", "fraction_ai", "Kategorie", "headline", "volltext"]
    for i, c in enumerate(vcols, 1):
        cell = ws3.cell(1, i, c)
        cell.font = HEAD
        cell.fill = FILL_HEAD
    rowi = 2
    for r in rows_sorted:
        fa = r["fraction_ai"]
        ws3.cell(rowi, 1, r["doc_id"])
        ws3.cell(rowi, 2, r.get("source", ""))
        ws3.cell(rowi, 3, r.get("pub_date", ""))
        ws3.cell(rowi, 4, round(fa, 4) if fa is not None else "n/a")
        ws3.cell(rowi, 5, cat_of(fa))
        ws3.cell(rowi, 6, r.get("headline", ""))
        ws3.cell(rowi, 7, r.get("text", ""))
        ws3.cell(rowi, 7).alignment = Alignment(wrap_text=True, vertical="top")
        f = fill_for(fa)
        if f:
            for ci in range(1, len(vcols) + 1):
                ws3.cell(rowi, ci).fill = f
        ws3.row_dimensions[rowi].height = 200
        rowi += 1
    for i, w in enumerate([36, 22, 12, 10, 14, 50, 100], 1):
        ws3.column_dimensions[get_column_letter(i)].width = w
    ws3.freeze_panes = "A2"

    # ============ Sheet 4: Worst-FP (für Manual Inspection) ============
    ws4 = wb.create_sheet("Worst-FP")
    ws4["A1"] = "Top 20 verdächtigste pre-2022 Texte (höchste fraction_ai = wahrscheinlichste False-Positives)"
    ws4["A1"].font = Font(bold=True, size=12)
    wcols = ["#", "fraction_ai", "Kategorie", "source", "pub_date", "chars", "headline",
             "ressort", "autor", "vorspann/volltext-Start"]
    for i, c in enumerate(wcols, 1):
        cell = ws4.cell(3, i, c)
        cell.font = HEAD
        cell.fill = FILL_HEAD
    rowi = 4
    for i, r in enumerate(rows_sorted[:20], 1):
        fa = r["fraction_ai"]
        ws4.cell(rowi, 1, i)
        ws4.cell(rowi, 2, round(fa, 4))
        ws4.cell(rowi, 3, cat_of(fa))
        ws4.cell(rowi, 4, r.get("source", ""))
        ws4.cell(rowi, 5, r.get("pub_date", ""))
        ws4.cell(rowi, 6, r.get("chars", 0))
        ws4.cell(rowi, 7, r.get("headline", ""))
        ws4.cell(rowi, 8, r.get("ressort", ""))
        ws4.cell(rowi, 9, r.get("autor", ""))
        ws4.cell(rowi, 10, (r.get("text") or "")[:600] + " …")
        ws4.cell(rowi, 10).alignment = Alignment(wrap_text=True, vertical="top")
        ws4.row_dimensions[rowi].height = 80
        f = fill_for(fa)
        if f:
            for ci in range(1, len(wcols) + 1):
                ws4.cell(rowi, ci).fill = f
        rowi += 1
    for i, w in enumerate([4, 10, 12, 22, 12, 8, 40, 20, 20, 80], 1):
        ws4.column_dimensions[get_column_letter(i)].width = w

    wb.save(OUT_XLSX)
    print(f"--- Excel: {OUT_XLSX}")
    print(f"--- {len(ok)} Docs, FPR-Strict (>= 0.5): {fp_strict}/{len(ok)} = "
          f"{fp_strict / max(len(ok), 1) * 100:.1f} %")


if __name__ == "__main__":
    main()
